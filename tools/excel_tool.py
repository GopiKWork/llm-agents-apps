"""
Excel Tool - Strands module-based tool for reading and analyzing Excel files.
Includes all Excel I/O helpers (openpyxl for reading, xlsxwriter for writing).
"""

import io
import openpyxl
import boto3
from typing import Any, Dict, List, Optional


# ============================================================================
# Excel I/O helpers
# ============================================================================

def read_workbook(filepath: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(filepath)


def read_workbook_from_s3(s3_path: str, aws_region: str = 'us-west-2') -> openpyxl.Workbook:
    if not s3_path.startswith('s3://'):
        raise ValueError("S3 path must start with 's3://'")
    parts = s3_path[5:].split('/', 1)
    bucket, key = parts[0], parts[1] if len(parts) > 1 else ''
    body = boto3.client('s3', region_name=aws_region).get_object(Bucket=bucket, Key=key)['Body'].read()
    return openpyxl.load_workbook(io.BytesIO(body))


def get_worksheet_names(workbook: openpyxl.Workbook) -> List[str]:
    return workbook.sheetnames


def read_worksheet_data(worksheet, max_rows: Optional[int] = None) -> List[List[Any]]:
    return [list(row) for row in worksheet.iter_rows(max_row=max_rows, values_only=True)]


def read_cell_value(worksheet, cell_ref: str) -> Any:
    return worksheet[cell_ref].value


def read_range_data(worksheet, start_cell: str, end_cell: str) -> List[List[Any]]:
    return [[cell.value for cell in row] for row in worksheet[start_cell:end_cell]]


def extract_images(worksheet) -> List[Dict[str, Any]]:
    images = []
    if hasattr(worksheet, '_images') and worksheet._images:
        for img in worksheet._images:
            images.append({
                'data': img._data(),
                'anchor': str(img.anchor) if hasattr(img, 'anchor') else None,
                'format': img.format if hasattr(img, 'format') else 'unknown'
            })
    return images


def extract_charts(workbook: openpyxl.Workbook, worksheet_name: str) -> List[Dict[str, Any]]:
    charts = []
    if worksheet_name not in workbook.sheetnames:
        return charts
    ws = workbook[worksheet_name]
    if hasattr(ws, '_charts') and ws._charts:
        for chart in ws._charts:
            title_text = None
            try:
                if hasattr(chart, 'title') and chart.title and hasattr(chart.title, 'tx') and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                        runs = chart.title.tx.rich.p[0].r
                        if runs:
                            title_text = runs[0].t
            except Exception:
                title_text = "Chart"
            position = None
            try:
                if hasattr(chart, 'anchor') and chart.anchor and hasattr(chart.anchor, '_from'):
                    position = f"{chr(65 + chart.anchor._from.col)}{chart.anchor._from.row + 1}"
            except Exception:
                position = "Unknown"
            charts.append({'type': chart.__class__.__name__, 'title': title_text, 'position': position})
    return charts


def save_image_bytes(image_bytes: bytes, output_path: str) -> None:
    with open(output_path, 'wb') as f:
        f.write(image_bytes)


# ============================================================================
# Helper to open a workbook and resolve sheet
# ============================================================================

def _open_sheet(tool_input, tool_use_id, from_s3=False):
    """Open workbook and resolve target sheet. Returns (workbook, worksheet, sheet_name) or error dict."""
    if from_s3:
        s3_path = tool_input.get("s3_path")
        if not s3_path:
            return _error(tool_use_id, "Error: s3_path is required")
        workbook = read_workbook_from_s3(s3_path, tool_input.get("aws_region", "us-west-2"))
    else:
        filepath = tool_input.get("filepath")
        if not filepath:
            return _error(tool_use_id, "Error: filepath is required")
        workbook = read_workbook(filepath)

    names = get_worksheet_names(workbook)
    if not names:
        return _error(tool_use_id, "Error: No sheets found in Excel file")

    target = tool_input.get("sheet_name") or names[0]
    if target not in names:
        return _error(tool_use_id, f"Error: Sheet '{target}' not found. Available: {', '.join(names)}")

    return workbook, workbook[target], target


def _ok(tid: str, text: str) -> dict:
    return {"toolUseId": tid, "status": "success", "content": [{"text": text}]}


def _error(tid: str, text: str) -> dict:
    return {"toolUseId": tid, "status": "error", "content": [{"text": text}]}


def _rows_to_csv(data):
    return "".join(",".join(str(c) if c is not None else "" for c in row) + "\n" for row in data)


# ============================================================================
# TOOL_SPEC and entry point
# ============================================================================

TOOL_SPEC = {
    "name": "excel_tool",
    "description": (
        "Excel file tool for reading and analyzing Excel files.\n\n"
        "Actions:\n"
        "- read_file: Read Excel file contents\n"
        "- read_s3: Read Excel file from S3\n"
        "- list_sheets: List all sheets\n"
        "- read_cell: Read specific cell value\n"
        "- read_range: Read cell range\n"
        "- extract_images: Extract embedded images\n"
        "- list_charts: List charts in a sheet\n"
        "- get_info: Get file information\n"
    ),
    "inputSchema": {
        "json": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "description": "Action to perform",
                    "enum": ["read_file", "read_s3", "list_sheets", "read_cell",
                             "read_range", "extract_images", "list_charts", "get_info"],
                },
                "filepath": {"type": "string", "description": "Path to Excel file"},
                "s3_path": {"type": "string", "description": "S3 path (s3://bucket/key)"},
                "sheet_name": {"type": "string", "description": "Sheet name (optional, defaults to first)"},
                "cell_ref": {"type": "string", "description": "Cell reference e.g. 'A1'"},
                "start_cell": {"type": "string", "description": "Start cell for read_range"},
                "end_cell": {"type": "string", "description": "End cell for read_range"},
                "output_dir": {"type": "string", "description": "Output directory for extract_images"},
                "aws_region": {"type": "string", "description": "AWS region (default: us-west-2)"},
            },
            "required": ["action"],
        }
    },
}


def excel_tool(tool: dict, **kwargs: Any) -> dict:
    """Excel file tool for reading and analyzing Excel files."""
    try:
        tid = tool.get("toolUseId", "default-id")
        inp = tool.get("input", {})
        action = inp.get("action")
        if not action:
            return _error(tid, "Error: action parameter is required")

        if action in ("read_file", "read_s3"):
            result = _open_sheet(inp, tid, from_s3=(action == "read_s3"))
            if isinstance(result, dict):
                return result
            _, ws, sheet = result
            data = read_worksheet_data(ws)
            if not data:
                return _error(tid, "Error: No data found in sheet")
            src = " (from S3)" if action == "read_s3" else ""
            return _ok(tid, f"Sheet: {sheet}{src}\nRows: {len(data)}\n\n{_rows_to_csv(data)}")

        elif action == "list_sheets":
            fp = inp.get("filepath")
            if not fp:
                return _error(tid, "Error: filepath is required")
            names = get_worksheet_names(read_workbook(fp))
            return _ok(tid, ", ".join(names) if names else "No sheets found")

        elif action == "read_cell":
            for req in ("filepath", "sheet_name", "cell_ref"):
                if not inp.get(req):
                    return _error(tid, f"Error: {req} is required for read_cell")
            wb = read_workbook(inp["filepath"])
            if inp["sheet_name"] not in get_worksheet_names(wb):
                return _error(tid, f"Error: Sheet '{inp['sheet_name']}' not found")
            val = read_cell_value(wb[inp["sheet_name"]], inp["cell_ref"])
            return _ok(tid, str(val) if val is not None else "Empty cell")

        elif action == "read_range":
            for req in ("filepath", "sheet_name", "start_cell", "end_cell"):
                if not inp.get(req):
                    return _error(tid, f"Error: {req} is required for read_range")
            wb = read_workbook(inp["filepath"])
            if inp["sheet_name"] not in get_worksheet_names(wb):
                return _error(tid, f"Error: Sheet '{inp['sheet_name']}' not found")
            data = read_range_data(wb[inp["sheet_name"]], inp["start_cell"], inp["end_cell"])
            header = f"Range {inp['start_cell']}:{inp['end_cell']} from {inp['sheet_name']}\n\n"
            return _ok(tid, header + _rows_to_csv(data))

        elif action == "extract_images":
            for req in ("filepath", "sheet_name"):
                if not inp.get(req):
                    return _error(tid, f"Error: {req} is required for extract_images")
            wb = read_workbook(inp["filepath"])
            sn = inp["sheet_name"]
            if sn not in get_worksheet_names(wb):
                return _error(tid, f"Error: Sheet '{sn}' not found")
            imgs = extract_images(wb[sn])
            if not imgs:
                return _ok(tid, f"No images found in sheet '{sn}'")
            out_dir = inp.get("output_dir", ".")
            lines = [f"Found {len(imgs)} image(s) in '{sn}':"]
            for i, img in enumerate(imgs, 1):
                path = f"{out_dir}/image_{sn}_{i}.{img['format']}"
                save_image_bytes(img['data'], path)
                lines.append(f"  {i}. Saved to {path} ({len(img['data'])} bytes, format: {img['format']})")
            return _ok(tid, "\n".join(lines))

        elif action == "list_charts":
            for req in ("filepath", "sheet_name"):
                if not inp.get(req):
                    return _error(tid, f"Error: {req} is required for list_charts")
            wb = read_workbook(inp["filepath"])
            sn = inp["sheet_name"]
            if sn not in get_worksheet_names(wb):
                return _error(tid, f"Error: Sheet '{sn}' not found")
            ch = extract_charts(wb, sn)
            if not ch:
                return _ok(tid, f"No charts found in sheet '{sn}'")
            lines = [f"Found {len(ch)} chart(s) in '{sn}':"]
            for i, c in enumerate(ch, 1):
                lines.append(f"  {i}. {c['type']} - '{c['title']}' at {c['position']}")
            return _ok(tid, "\n".join(lines))

        elif action == "get_info":
            fp = inp.get("filepath")
            if not fp:
                return _error(tid, "Error: filepath is required")
            wb = read_workbook(fp)
            names = get_worksheet_names(wb)
            lines = [f"File: {fp}", f"Sheets: {len(names)}", ""]
            for sn in names:
                ws = wb[sn]
                lines.append(f"Sheet: {sn}")
                lines.append(f"  Rows: {ws.max_row}")
                lines.append(f"  Columns: {ws.max_column}")
                ni = len(extract_images(ws))
                nc = len(extract_charts(wb, sn))
                if ni:
                    lines.append(f"  Images: {ni}")
                if nc:
                    lines.append(f"  Charts: {nc}")
                lines.append("")
            return _ok(tid, "\n".join(lines))

        else:
            return _error(tid, f"Error: Unknown action '{action}'")

    except Exception as e:
        return _error(tool.get("toolUseId", "default-id"), f"Error: {str(e)}")
